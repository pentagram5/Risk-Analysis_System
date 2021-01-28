# 패키지 로드
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import pandas as pd
import datetime as dt
import folium
from folium.plugins import HeatMap
import io
from PIL import Image
import json
import requests
from functools import reduce

import chart_studio.plotly as py
import plotly.express as px
import plotly.graph_objs as go

import mapboxgl
from mapboxgl.viz import *    
from mapboxgl.utils import create_color_stops 
from mapboxgl.utils import create_numeric_stops
from dateutil.relativedelta import relativedelta
import branca.colormap as cmp
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image
from openpyxl.utils.dataframe import dataframe_to_rows


# 한글 패치
font_fname = '/usr/share/fonts/truetype/nanum/NanumGothic.ttf'
font_name = fm.FontProperties(fname=font_fname).get_name()
matplotlib.rc('font', family=font_name)
fm._rebuild()

    
# 지역별 민원
df = pd.read_excel('db_v0.7.xlsx', sheet_name='Database_1')
df_dong_list = pd.read_excel('db_v0.7.xlsx', sheet_name='dong_list')

# 구별, 동별 하수관로
df_pipe = pd.read_excel('db_v0.7.xlsx', sheet_name='Database_4', skiprows = 1)
df_pipe_total = df_pipe.iloc[:,0:6]
df_pipe_add = df_pipe.loc[:,['구','2014.1','2015.1','2016.1','2017.1','2018.1']]
df_pipe_add.columns = ['구','2014','2015','2016','2017','2018']

# 구 전체
gu_list = list(df['유형_소'].unique())
gu_list.sort()

seoul_dict={}
for i in gu_list:
    temp = list(df[df['유형_소']==i]['동_수정'].unique())
    temp.sort()
    temp.insert(0, '전체')
    seoul_dict[i]=temp
    
    
# 4. 하수관로 연장 및 준설대비 악취발생 분석
### 4.1.1. 자치구별 민원발생건수/총 하수관로연장
# 구별 통계 산출 - 자치구별 민원발생건수/총 하수관로연장
def get_pipe_total_stat_gu(gu, start_date, end_date) : 
  global df_output
  df_fin = df[(df['datetime'] > start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  df_output = pd.DataFrame(df_fin[df_fin['유형_소'] == gu]['유형_소'].groupby(df_fin['y-m']).count()).reset_index()
  df_output.columns = ['연월', '민원 빈도수']
  df_output['연도별 총하수관로'] = ''

  for i in range(len(df_output)) :
    YM = int(df_output['연월'][i].split('-')[0])
    df_output['연도별 총하수관로'][i] = int(df_pipe_total[df_pipe_total['구'] == gu][YM])

  name = gu
  df_output[name] = ((df_output['민원 빈도수']/df_output['연도별 총하수관로'])).astype(float)

  for i in range(len(df_output)) :
    #df_output[name][i] = '%0.6f'%df_output[name][i]
    df_output[name][i] = round(df_output[name][i],6)*1000000
  df_output = df_output[['연월', name]].fillna(0)
  return print('생성',gu, start_date, end_date)


# 구 1개 이상 선택시 통계 산출
def get_pipe_total_stat_gus(gu_list, start_date, end_date,tot_num) : 
  global df_output
  get_pipe_total_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_total_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_total_stat_gu_list.append(df_output)

  df_output = reduce(lambda left, right: pd.merge(left, right, on='연월', how = 'outer'), get_pipe_total_stat_gu_list)
  df_output = df_output.sort_values(by = '연월').fillna(0)
  get_pipe_stat_gus_sum(gu_list, df_output, tot_num) 
  write_to_html_file(df_output, tot_num, 'templates/df{}.html'.format(tot_num))
  return print('생성')


# 구 1개 이상 선택시 그래프 산출
def get_pipe_total_graph_gus(gu_list, start_date, end_date, tot_num) : 
  get_pipe_total_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_total_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_total_stat_gu_list.append(df_output)
  if len(get_pipe_total_stat_gu_list) > 1 : 
    #############################################
    trace_list = []
    for i in range(len(get_pipe_total_stat_gu_list)) : 
      trace_list.append(go.Scatter(x = get_pipe_total_stat_gu_list[i][get_pipe_total_stat_gu_list[i].columns[0]], y = get_pipe_total_stat_gu_list[i][get_pipe_total_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(건/월)")
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.write_html('templates/graph_{}.html'.format(tot_num)) 
    #############################################
    df_merge_pre = get_pipe_total_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(x ='연월', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(건/월)')
    ax.title.set_text('<구별 악취민원발생수>')  
    plt.savefig('templates/df{}.jpg'.format(tot_num))  
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num), sheet_name = 'Data') 
    #############################################
    return print('생성')

  if len(get_pipe_total_stat_gu_list) == 1 : 
    #############################################
    trace_list = []
    for i in range(len(get_pipe_total_stat_gu_list)) : 
      trace_list.append(go.Bar(x = get_pipe_total_stat_gu_list[i][get_pipe_total_stat_gu_list[i].columns[0]], y = get_pipe_total_stat_gu_list[i][get_pipe_total_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_traces(text = trace_list[0]['y'],  textposition='outside')
    fig.update_traces(marker_color='rgb(0,0,0)')
    fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(건/월)")
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.write_html('templates/graph_{}.html'.format(tot_num)) 
    #############################################
    df_merge_pre = get_pipe_total_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(kind = 'bar', x ='연월', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(건/월)')
    ax.title.set_text('<구별 악취민원발생수>')  
    plt.savefig('templates/df{}.jpg'.format(tot_num))  
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num), sheet_name='Data') 
    #############################################
    return print('생성')


# 구별 지도 산출 - 자치구별 민원발생건수/총 하수관로연장
def get_pipe_total_map_by_gu(gu_list, start_date, end_date, tot_num) :  
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  df_output = pd.DataFrame(df_fin.groupby(df_fin['유형_소']).count()).reset_index()[['유형_소','번호']]

  df_output.columns = ['구', '민원수']
  YM = int(end_date.split('-')[0])

  df_pipe_total_gu = df_pipe_total[['구',YM]].reset_index(drop=True).sort_values(by = '구').reset_index(drop=True)
  df_pipe_total_gu_fin = pd.merge(df_output, df_pipe_total_gu, on = '구', how='left')
  df_pipe_total_gu_fin['총 하수관로연장당 민원수'] = df_pipe_total_gu_fin['민원수']/df_pipe_total_gu_fin[YM]

  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
      data = json.loads(f.read())
  
  df_pipe_total_gu_fin = df_pipe_total_gu_fin[df_pipe_total_gu_fin['구'].str.contains('|'.join(gu_list))].reset_index(drop=True) ### 201219 추가
    
  data_list = []
  for i in range(len(data['features'])) : 
    data['features'][i]['properties']['자치구명'] = data['features'][i]['properties'].pop("name")
    for j in range(len(df_pipe_total_gu_fin)) : 
      if data['features'][i]['properties']["자치구명"] == df_pipe_total_gu_fin["구"][j] : 
        del data['features'][i]['properties']['code']
        del data['features'][i]['properties']['base_year']
        del data['features'][i]['properties']['name_eng']
        data['features'][i]['properties']["총 하수관로연장당 민원수"] = round(df_pipe_total_gu_fin["총 하수관로연장당 민원수"][j], 5)
        data['features'][i]['properties']["분석기간"] =  start_date + "~" + end_date
        data_list.append(data['features'][i])

  data['features'] = data_list  
  token = 'pk.eyJ1IjoiZHVjazk2NjciLCJhIjoiY2thNzRiMXhxMGQ5NTJ0cXB0dGpmZ3RrOSJ9.f45On_lrv6Nm4iO7oCg7nw'
  res, bins = pd.qcut(df_pipe_total_gu_fin['총 하수관로연장당 민원수'], 3, retbins=True)

  color_breaks = color_breaks = list(bins) 
  color_stops = create_color_stops(color_breaks, colors='BuPu')

  viz = ChoroplethViz(access_token=token, data=data, color_property='총 하수관로연장당 민원수', color_stops=color_stops, center = [126.986, 37.565] , zoom=9.5, line_width=2)
  viz.create_html('templates/map_{}.html'.format(tot_num))
  return print('생성')      


### 4.1.2. 자치구별 민원발생건수/하수준설관로연장
# 구별 통계 산출 - 자치구별 민원발생건수/하수준설관로연장
def get_pipe_add_stat_gu(gu, start_date, end_date) : 
  global df_output
  df_fin = df[(df['datetime'] > start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  df_output = pd.DataFrame(df_fin[df_fin['유형_소'] == gu]['유형_소'].groupby(df_fin['y-m']).count()).reset_index()
  df_output.columns = ['연월', '민원 빈도수']
  df_output['연도별 하수준설관로'] = ''

  # 특정구, 특정연도 총 하수관로
  for i in range(len(df_output)) :
    YM = df_output['연월'][i].split('-')[0]
    df_output['연도별 하수준설관로'][i] = int(df_pipe_add[df_pipe_add['구'] == gu][YM])

  name = gu
  df_output[name] = ((df_output['민원 빈도수']/df_output['연도별 하수준설관로'])).astype(float)

  for i in range(len(df_output)) :
    df_output[name][i] = round(df_output[name][i],6)*1000000
  df_output = df_output[['연월', name]].fillna(0)
  return print('생성') 


# 구 1개 이상 선택시 통계 산출
def get_pipe_add_stat_gus(gu_list, start_date, end_date,tot_num) : 
  global df_output
  get_pipe_add_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_add_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_add_stat_gu_list.append(df_output)

  df_output = reduce(lambda left, right: pd.merge(left, right, on='연월', how = 'outer'), get_pipe_add_stat_gu_list)
  df_output = df_output.sort_values(by = '연월').fillna(0)
  get_pipe_stat_gus_sum(gu_list, df_output, tot_num) # 최대 최소 추가
  write_to_html_file(df_output, tot_num, 'templates/df{}.html'.format(tot_num))
  return print('생성')


# 구 1개 이상 선택시 그래프 산출
def get_pipe_add_graph_gus(gu_list, start_date, end_date, tot_num) : 
  get_pipe_add_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_add_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_add_stat_gu_list.append(df_output)

  if len(get_pipe_add_stat_gu_list) > 1 :     
    #############################################  
    trace_list = []
    for i in range(len(get_pipe_add_stat_gu_list)) : 
      trace_list.append(go.Scatter(x = get_pipe_add_stat_gu_list[i][get_pipe_add_stat_gu_list[i].columns[0]], y = get_pipe_add_stat_gu_list[i][get_pipe_add_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.update_layout(title = {'text': "<동별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(연/월)")
    fig.write_html('templates/graph_{}.html'.format(tot_num))
    #############################################  
    df_merge_pre = get_pipe_add_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(x ='연월', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(연/월)')
    ax.title.set_text('<구별 악취민원발생수>')
    plt.savefig('templates/df{}.jpg'.format(tot_num)) 
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num), sheet_name ="Data")   
    #############################################
    return print('생성')

  if len(get_pipe_add_stat_gu_list) == 1 :     
    #############################################  
    trace_list = []
    for i in range(len(get_pipe_add_stat_gu_list)) : 
      trace_list.append(go.Scatter(x = get_pipe_add_stat_gu_list[i][get_pipe_add_stat_gu_list[i].columns[0]], y = get_pipe_add_stat_gu_list[i][get_pipe_add_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_traces(text = trace_list[0]['y'],  textposition='outside')
    fig.update_traces(marker_color='rgb(0,0,0)')
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(연/월)")
    fig.write_html('templates/graph_{}.html'.format(tot_num))
    #############################################  
    df_merge_pre = get_pipe_add_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(kind = 'bar', x ='연월', figsize = (18,12), rot = 45, color = 'black')
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(연/월)')
    ax.title.set_text('<구별 악취민원발생수>')
    plt.savefig('templates/df{}.jpg'.format(tot_num)) 
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num))   
    #############################################
    return print('생성')


# 구별 지도 산출 - 자치구별 민원발생건수/하수준설관로연장
def get_pipe_add_map_by_gu(gu_list, start_date, end_date, tot_num) :  
  df_fin = df[(df['datetime'] > start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  df_output = pd.DataFrame(df_fin.groupby(df_fin['유형_소']).count()).reset_index()[['유형_소','번호']]

  df_output.columns = ['구', '민원수']
  YM = end_date.split('-')[0]

  df_pipe_add_gu = df_pipe_add[['구',YM]].reset_index(drop=True).sort_values(by = '구').reset_index(drop=True)
  df_pipe_add_gu_fin = pd.merge(df_output, df_pipe_add_gu, on = '구', how='left' )
  df_pipe_add_gu_fin['하수준설관로연장 민원수'] = df_pipe_add_gu_fin['민원수']/df_pipe_add_gu_fin[YM]

  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
      data = json.loads(f.read())
        
  df_pipe_add_gu_fin = df_pipe_add_gu_fin[df_pipe_add_gu_fin['구'].str.contains('|'.join(gu_list))].reset_index(drop=True) ### 201219 추가

  data_list = []
  for i in range(len(data['features'])) : 
    data['features'][i]['properties']['자치구명'] = data['features'][i]['properties'].pop("name")
    for j in range(len(df_pipe_add_gu_fin)) : 
      if data['features'][i]['properties']["자치구명"] == df_pipe_add_gu_fin["구"][j] : 
            del data['features'][i]['properties']['code']
            del data['features'][i]['properties']['base_year']
            del data['features'][i]['properties']['name_eng']
            data['features'][i]['properties']["하수준설관로연장 민원수"] = round(df_pipe_add_gu_fin["하수준설관로연장 민원수"][j], 5)
            data['features'][i]['properties']["분석기간"] =  start_date + "~" + end_date
            data_list.append(data['features'][i])
  
  data['features'] = data_list 
  token = 'pk.eyJ1IjoiZHVjazk2NjciLCJhIjoiY2thNzRiMXhxMGQ5NTJ0cXB0dGpmZ3RrOSJ9.f45On_lrv6Nm4iO7oCg7nw'
  res, bins = pd.qcut(df_pipe_add_gu_fin['하수준설관로연장 민원수'], 5, retbins=True)
  color_breaks = color_breaks = list(bins) 
  color_stops = create_color_stops(color_breaks, colors='BuPu')

  viz = ChoroplethViz(access_token=token, data=data, color_property= '하수준설관로연장 민원수', color_stops=color_stops, center = [126.986, 37.565], zoom=9.5, line_width=2)
  viz.create_html('templates/map_{}.html'.format(tot_num))
  return print('생성')    


### 4.1.3. (준설관로연장/총 하수관로연장) * 100(%)
# 구별 통계 산출 - (준설관로연장/총 하수관로연장) * 100(%)
def get_pipe_total_add_stat_gu(gu, start_date, end_date) : 
  global df_output
  df_pipe_add_fin = df_pipe_add[df_pipe_add['구'] == gu].reset_index(drop=True).iloc[:,1:6]
  df_pipe_total_fin = df_pipe_total[df_pipe_total['구'] == gu].reset_index(drop=True).iloc[:,1:6]

  df_total_add = pd.DataFrame(df_pipe_add_fin.values[0] / df_pipe_total_fin.values[0]).transpose() * 100
  df_total_add = df_total_add.transpose()
  df_total_add.index = df_pipe_add.columns[1:6]
  df_total_add.index.name = '연월'
  df_total_add.columns = [gu]
  df_output = df_total_add.reset_index()
  return print('생성',gu, start_date, end_date)


# 구 1개 이상 선택시 통계 산출
def get_pipe_total_add_stat_gus(gu_list, start_date, end_date, tot_num) : 
  global df_output
  get_pipe_total_add_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_total_add_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_total_add_stat_gu_list.append(df_output)

  df_output = reduce(lambda left, right: pd.merge(left, right, on='연월', how = 'outer'), get_pipe_total_add_stat_gu_list)
  get_pipe_stat_gus_sum(gu_list, df_output, tot_num) 
  write_to_html_file(df_output, tot_num, 'templates/df{}.html'.format(tot_num))
  return print('생성')


# 구 1개 이상 선택시 그래프 산출
def get_pipe_total_add_graph_gus(gu_list, start_date, end_date, tot_num) :
  get_pipe_total_add_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_pipe_total_add_stat_gu(gu_list[i], start_date, end_date)
    get_pipe_total_add_stat_gu_list.append(df_output)
  if len(get_pipe_total_add_stat_gu_list) > 1 : 
    #############################################
    trace_list = []
    for i in range(len(get_pipe_total_add_stat_gu_list)) : 
      trace_list.append(go.Scatter(x = get_pipe_total_add_stat_gu_list[i][get_pipe_total_add_stat_gu_list[i].columns[0]], y = get_pipe_total_add_stat_gu_list[i][get_pipe_total_add_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(건/월)")
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.write_html('templates/graph_{}.html'.format(tot_num)) 
    #############################################
    df_merge_pre = get_pipe_total_add_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(x ='연월', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(건/월)')
    ax.title.set_text('<구별 악취민원발생수>')  
    plt.savefig('templates/df{}.jpg'.format(tot_num))  
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num), sheet_name='Data')
    #############################################
    return print('생성')

  if len(get_pipe_total_add_stat_gu_list) == 1 : 
    #############################################
    trace_list = []
    for i in range(len(get_pipe_total_add_stat_gu_list)) : 
      trace_list.append(go.Bar(x = get_pipe_total_add_stat_gu_list[i][get_pipe_total_add_stat_gu_list[i].columns[0]], y = get_pipe_total_add_stat_gu_list[i][get_pipe_total_add_stat_gu_list[i].columns[1]],  name = gu_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원 건수(건/월)")
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.write_html('templates/graph_{}.html'.format(tot_num)) 
    #############################################
    df_merge_pre = get_pipe_total_add_stat_gus(gu_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '연월').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(kind = 'bar', x ='연월', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(건/월)')
    ax.title.set_text('<구별 악취민원발생수>')  
    plt.savefig('templates/df{}.jpg'.format(tot_num))  
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num))
    #############################################
    return print('생성')


# 구별 지도 산출 - (준설관로연장/총 하수관로연장) * 100(%)
def get_pipe_total_add_map_by_gu(gu_list, year, tot_num,  start_date, end_date) :  
  df_pipe_total_add_gu_fin = get_pipe_total_add_stat_gus(gu_list, start_date, end_date, tot_num)
  df_pipe_total_add_gu_fin = df_output.transpose()
  df_pipe_total_add_gu_fin = df_pipe_total_add_gu_fin.rename(columns=df_pipe_total_add_gu_fin.iloc[0])
  df_pipe_total_add_gu_fin = df_pipe_total_add_gu_fin.drop(df_pipe_total_add_gu_fin.index[0])
  
  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
      data = json.loads(f.read())
  
  data_list = []
  for i in range(len(data['features'])) : 
    data['features'][i]['properties']['자치구명'] = data['features'][i]['properties'].pop("name")
    for j in range(len(df_pipe_total_add_gu_fin)) : 
      if data['features'][i]['properties']["자치구명"] == df_pipe_total_add_gu_fin.index[j] : 
        del data['features'][i]['properties']['code']
        del data['features'][i]['properties']['base_year']
        del data['features'][i]['properties']['name_eng']
        data['features'][i]['properties']["비율"] = round(df_pipe_total_add_gu_fin[year][j], 5)
        data['features'][i]['properties']["분석기간"] =  str(year) + "년"
        data_list.append(data['features'][i])

  data['features'] = data_list  
  token = 'pk.eyJ1IjoiZHVjazk2NjciLCJhIjoiY2thNzRiMXhxMGQ5NTJ0cXB0dGpmZ3RrOSJ9.f45On_lrv6Nm4iO7oCg7nw'
  res, bins = pd.qcut(df_pipe_total_add_gu_fin[year], 3, retbins=True)
  color_breaks = color_breaks = list(bins) 
  color_stops = create_color_stops(color_breaks, colors='BuPu')

  viz = ChoroplethViz(access_token=token, data=data, color_property= "비율", color_stops=color_stops, center = [126.986, 37.565] , zoom=9.5, line_width=2)
  viz.create_html('templates/map_{}.html'.format(tot_num))
  return print('생성')   


#평균, 최대 최소 출력용
def get_pipe_stat_gus_sum(gu_list, df_output, tot_num):
    df_temp = pd.DataFrame()
    for i in gu_list:
        temp_list =[]
        temp_list.append(round(df_output[i].mean(), 2))
        if df_output[i].max()==0:
            temp_list.append(str(0))
        elif df_output[i].empty:
            temp_list.append(str(0))
        else: temp_list.append(str(df_output[df_output[i] == df_output[i].max()]['연월'].values[0]) + '  값:' + str(round(df_output[i].max(), 2)))
            
        if len(df_output[df_output[i]==0]) >1:
            temp_list.append(str(0))
        elif df_output[i].empty:
            temp_list.append(str(0))
        else: temp_list.append(str(df_output[df_output[i] == df_output[i].min()]['연월'].values[0]) +  '  값:' + str(round(df_output[i].min(),2)))
                         
        df_temp[i] = temp_list
    df_temp.index = ['평균값', '최대값', '최소값']
    df_temp.to_excel('templates/df_sum{}.xlsx'.format(tot_num))
    write_to_html_file(df_temp, tot_num, 'templates/df_sum{}.html'.format(tot_num))
    return print('생성')


def write_to_html_file(df, tot_num, filename='out.html' ):
       '''
       Write an entire dataframe to an HTML file with nice formatting.
       '''

       result = '''
<html>
<head>
<style>

       h2 {
           text-align: center;
           font-family: Helvetica, Arial, sans-serif;
       }
       table { 
           margin-left: auto;
           margin-right: auto;
       }
       table, th, td {
           border: 1px solid black;
           border-collapse: collapse;
       }
       th, td {
           padding: 5px;
           text-align: center;
           font-family: Helvetica, Arial, sans-serif;
           font-size: 50%;
       }
       thead tr {
        background-color: black;
        color: white;
       }
        tbody tr:nth-child(2n) {
        background-color: #dcdcdc;
       }
        tbody tr:nth-child(2n+1) {
        background-color: #bebebe;
       }
       .wide {
           width: 90%; 
       }

</style>
</head>
<body>
       '''
   
       result += df.to_html(classes='wide', escape=False)
       result += '''
</body>
</html>
'''
       with open(filename, 'w') as f:
           f.write(result)
            
            
def get_pipe_export(gu_list, type, start_date, end_date, tot_num, year) :   
  if type == "total" :
    get_pipe_total_graph_gus(gu_list, start_date, end_date, tot_num)
    get_pipe_total_map_by_gu(gu_list, start_date, end_date, tot_num)

  elif type == "add" :
    get_pipe_add_graph_gus(gu_list, start_date, end_date, tot_num)
    get_pipe_add_map_by_gu(gu_list, start_date, end_date, tot_num)

  elif type == "total_add" :
    get_pipe_total_add_graph_gus(gu_list, start_date, end_date, tot_num)
    get_pipe_total_add_map_by_gu(gu_list, year, tot_num, start_date, end_date)
  
  df_sum = pd.read_excel('templates/df_sum{}.xlsx'.format(tot_num))
  wb = openpyxl.load_workbook("templates/df{}.xlsx".format(tot_num)) # 엑셀 파일 선택
  sheet_last = wb.create_sheet("Graph") # 제일 마지막에 시트 추가
  
  ws = wb.worksheets[1] # 시트 지정
  img = Image.open("templates/df{}.jpg".format(tot_num)) # 이미지 지정
  img = openpyxl.drawing.image.Image("templates/df{}.jpg".format(tot_num)) # 이미지 지정
  ws.add_image(img,'A1') # 이미지 삽입

  sheet_last = wb.create_sheet("Sum") # 제일 마지막에 시트 추가
  ws = wb.worksheets[2] # 시트 지정
  for row in dataframe_to_rows(df_sum, index = True, header= True) :
    if len(row) > 1 :
      ws.append(row)

  wb.save("templates/df{}.xlsx".format(tot_num)) # 엑셀 저장