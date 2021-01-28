# 패키지 로드
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import pandas as pd
import datetime as dt
import folium
from folium.plugins import HeatMap
import io
from PIL import Image
import json
import requests
from functools import reduce

import chart_studio.plotly as py
import plotly.express as px
import plotly.graph_objs as go

import mapboxgl
from mapboxgl.viz import *
from mapboxgl.utils import create_color_stops 
from mapboxgl.utils import create_numeric_stops
from dateutil.relativedelta import relativedelta
import branca.colormap as cmp
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from PIL import Image


# 한글 패치
font_fname = '/usr/share/fonts/truetype/nanum/NanumGothic.ttf'
font_name = fm.FontProperties(fname=font_fname).get_name()
matplotlib.rc('font', family=font_name)
fm._rebuild()

    
# 데이터 로드 지역별 민원
df = pd.read_excel('db_v0.7.xlsx', sheet_name='Database_1')
df_dong_list = pd.read_excel('db_v0.7.xlsx', sheet_name='dong_list')

# 구 전체
gu_list = list(df['유형_소'].unique())
gu_list.sort()
seoul_dict={}
for i in gu_list:
        temp = list(df[df['유형_소']==i]['동_수정'].unique())
        temp.sort()
        temp.insert(0, '전체')
        seoul_dict[i]=temp

        
def get_dict():
    gu_list = list(df['유형_소'].unique())
    gu_list.sort()
    gu_list.insert(0,'서울시 전체')
    seoul_dict={}
    for i in gu_list:
        temp = list(df[df['유형_소']==i]['동_수정'].unique())
        temp.sort()
        temp.insert(0, '전체')
        seoul_dict[i]=temp
    return seoul_dict

def get_dict_2():
    gu_list = list(df['유형_소'].unique())
    gu_list.sort()
    gu_list.insert(0,'전체')
    seoul_dict={}
    for i in gu_list:
        temp = list(df[df['유형_소']==i]['동_수정'].unique())
        temp.sort()
        temp.insert(0, '전체')
        seoul_dict[i]=temp
    return seoul_dict
        
# 구별 통계 출력
def get_stat_gu(gu, start_date, end_date) : 
  global df_output
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if len(df_fin[df_fin["유형_소"] == gu]) > 0 :
    df_output = pd.DataFrame(df_fin[df_fin['유형_소'] == gu]['유형_소'].groupby(df_fin['y-m']).count()).reset_index()
    df_output.columns = ['날짜', '{} 민원수(건)'.format(gu)]
    return print('생성',gu, start_date, end_date)
  else :
    return print("값없음")  


# 구 1개 이상 선택시 통계 산출
def get_stat_gus(gu_list, start_date, end_date, tot_num) : 
  global df_output
  get_stat_gu_list = []
  for i in range(len(gu_list)) :
    get_stat_gu(gu_list[i], start_date, end_date)
    get_stat_gu_list.append(df_output)
  df_output = reduce(lambda left, right: pd.merge(left, right, on='날짜', how = 'outer'), get_stat_gu_list)
  write_to_html_file(df_output, tot_num, 'templates/df{}.html'.format(tot_num))
  return print('생성')


# 구별 그래프
def get_graph_gus(gu_list, start_date, end_date, tot_num) : 
    get_stat_gu_list = []
    for i in range(len(gu_list)) :
      get_stat_gu(gu_list[i], start_date, end_date)
      get_stat_gu_list.append(df_output)

    if len(get_stat_gu_list) > 1 :      
      #############################################
      trace_list = []
      for i in range(len(get_stat_gu_list)) : 
        trace_list.append(go.Scatter(x = get_stat_gu_list[i][get_stat_gu_list[i].columns[0]], y = get_stat_gu_list[i][get_stat_gu_list[i].columns[1]],  name = gu_list[i]))
      fig = go.Figure(data = trace_list)
      fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원건수(건/월)")
      fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
      fig.write_html('templates/graph_{}.html'.format(tot_num))
      #############################################
      df_merge_pre = get_stat_gus(gu_list, start_date, end_date, tot_num)
      df_merge = df_output.sort_values(by = '날짜').fillna(0)
      ax = df_merge.plot(x ='날짜', figsize = (18,12))
      ax.set_xlabel('기간(연도/월)')
      ax.set_ylabel('악취민원 건수(건/월)')
      ax.title.set_text('<구별 악취민원발생수>')
      plt.savefig('templates/df{}.jpg'.format(tot_num))  
      df_output.to_excel('templates/df{}.xlsx'.format(tot_num),sheet_name='Data')
      #############################################
      return print('생성') 
  
    if len(get_stat_gu_list) == 1 :
      #############################################
      trace_list = []
      for i in range(len(get_stat_gu_list)) : 
        trace_list.append(go.Bar(x = get_stat_gu_list[i][get_stat_gu_list[i].columns[0]], y = get_stat_gu_list[i][get_stat_gu_list[i].columns[1]],  name = gu_list[i]))
      fig = go.Figure(data = trace_list)
      fig.update_traces(text = trace_list[0]['y'],  textposition='outside')
      fig.update_traces(marker_color='rgb(0,0,0)')
      fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
      fig.update_layout(title = {'text': "<구별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원건수(건/월)")
      fig.write_html('templates/graph_{}.html'.format(tot_num))
      #############################################
      df_merge_pre = get_stat_gus(gu_list, start_date, end_date, tot_num)
      df_merge = df_output.sort_values(by = '날짜').fillna(0)
      ax = df_merge.plot(kind = 'bar', x ='날짜', figsize = (18,12), rot = 45, color = 'black')
      ax.set_xlabel('기간(연도/월)')
      ax.set_ylabel('악취민원 건수(건/월)')
      ax.title.set_text('<구별 악취민원발생수>')
      plt.savefig('templates/df{}.jpg'.format(tot_num), sheet_name='Data')  
      df_output.to_excel('templates/df{}.xlsx'.format(tot_num))
      #############################################
      return print('생성')


# 구별, 히트맵 출력
def get_map_gu(gu, start_date, end_date, tot_num) :
  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
    data = json.loads(f.read())

  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if gu == '서울시' :    
    df_fin = df_fin.reset_index()
    latlon = []
    for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
    m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 11.5, tiles = 'Cartodb Positron')
    HeatMap(latlon, name = 'Heat Map').add_to(m)

    # 레이어 추가
    def piste_style_function(feature):
        return {'opacity': 1, 'weight': 1}
    folium.GeoJson(geo_data, style_function = piste_style_function, name='Boundary').add_to(m)
    
    # 버튼 추가
    folium.LayerControl().add_to(m)      
    
    # 범례
    linear = cmp.LinearColormap(['blue', 'green', 'yellow', 'red'], caption='Color Scale for Heat Map')
    linear.add_to(m)
    m.save('templates/heatmap{}.html'.format(tot_num))
    return print('생성')

  else :
    df_fin = df_fin[df_fin['유형_소'] == gu].reset_index()
    if len(df_fin[df_fin["유형_소"] == gu]) > 0 :
      latlon = []
      for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
      m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 13, tiles = 'Cartodb Positron')
      HeatMap(latlon).add_to(m)
      
      # 레이어 추가
      def piste_style_function(feature):
          return {'opacity': 1, 'weight': 1}
      folium.GeoJson(geo_data, style_function = piste_style_function, name='Boundary').add_to(m)
      
      # 버튼 추가
      folium.LayerControl().add_to(m)      
      
      # 범례
      linear = cmp.LinearColormap(['blue', 'green', 'yellow', 'red'], caption='Color Scale for Heat Map')
      linear.add_to(m)
      m.save('templates/heatmap{}.html'.format(tot_num))
      return print('생성')

    else :
      return print("값없음")  


# 구별, 포인트맵 출력
def get_map_point_gu(gu, start_date, end_date, tot_num) :
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if gu == '서울시' :    
    df_fin = df_fin.reset_index()
    latlon = []
    for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
    m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 11.5, tiles = 'Cartodb Positron')
    for i in range(len(latlon)) :
        folium.Circle(location = latlon[i], radius = 50, color = 'Blue', fill_color="#ccc", fill_opacity=1, weight=1, opacity=1, tooltip = df_fin.loc[i,'위반장소']).add_to(m) 
    m.save('templates/mappoint{}.html'.format(tot_num))
    return print("생성")  

  else :
    df_fin = df_fin[df_fin['유형_소'] == gu].reset_index()
    if len(df_fin[df_fin["유형_소"] == gu]) > 0 :
      latlon = []
      for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
      m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 13, tiles = 'Cartodb Positron')
      for i in range(len(latlon)) :
        folium.Circle(location = latlon[i], radius = 50, color = 'Blue', fill_color="#ccc", fill_opacity=1, weight=1, opacity=1, tooltip = df_fin.loc[i,'위반장소']).add_to(m) 
      m.save('templates/mappoint{}.html'.format(tot_num))
      return print("생성")  
    else :
      return print("값없음")  


# 전체 구 지도 비교
def get_map_compare_gu(start_date, end_date, tot_num) :  
  df_fin = df[(df['datetime'] > start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  df_output = pd.DataFrame(df_fin.groupby(df_fin['유형_소']).count()).reset_index()[['유형_소','번호']]
  df_output.columns = ['구', '민원수(건)']

  start_date = start_date.split("-")[0] + '년 ' + start_date.split("-")[1] + '월'
  end_date = end_date.split("-")[0] + '년 ' + end_date.split("-")[1] + '월'

  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
      data = json.loads(f.read())

  for i in range(len(data['features'])) : 
    data['features'][i]['properties']['자치구명'] = data['features'][i]['properties'].pop("name")
    for j in range(len(df_output)) : 
      if data['features'][i]['properties']["자치구명"] == df_output["구"][j] :
        del data['features'][i]['properties']['code']
        del data['features'][i]['properties']['base_year']
        del data['features'][i]['properties']['name_eng']
        data['features'][i]['properties']["분석기간"] =  start_date + "~" + end_date
        data['features'][i]['properties']["민원수(건)"] = int(df_output["민원수(건)"][j])  

  token = 'pk.eyJ1IjoiZHVjazk2NjciLCJhIjoiY2thNzRiMXhxMGQ5NTJ0cXB0dGpmZ3RrOSJ9.f45On_lrv6Nm4iO7oCg7nw'
  res, bins = pd.qcut(df_output['민원수(건)'], 3, retbins=True)
  color_breaks = list(bins) 
  color_stops = create_color_stops(color_breaks, colors='BuPu')

  viz = ChoroplethViz(access_token=token, data=data, color_property='민원수(건)', color_stops=color_stops, center=[126.986, 37.565], zoom=9.5, line_width=2)
  viz.create_html('templates/map_{}.html'.format(tot_num))
  return print("생성")


# 동별 출력
def get_stat_dong(gu, dong, start_date, end_date) : 
  global df_output
  df_fin = df[(df['datetime'] > start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if len(df_fin[df_fin["동_수정"] == dong]) > 0 :
    df_output = pd.DataFrame(df_fin[(df_fin['유형_소'] == gu)&(df_fin['동_수정'] == dong)]['동_수정'].groupby(df_fin['y-m']).count()).reset_index()
    df_output.columns = ['날짜', '{} 민원수(건)'.format(dong)]
    return print("생성") 
  else :
    return print("값없음")  


# 동 1개 이상 선택시 통계 산출
def get_stat_dongs(gu, dong_list, start_date, end_date, tot_num) : 
  global df_output
  get_stat_dong_list = []
  for i in range(len(dong_list)) :
    get_stat_dong(gu, dong_list[i], start_date, end_date)
    get_stat_dong_list.append(df_output)
  df_output = reduce(lambda left, right: pd.merge(left, right, on='날짜', how = 'outer'), get_stat_dong_list)
  df_output = df_output.sort_values(by = '날짜').fillna(0)
  write_to_html_file(df_output, tot_num, 'templates/df{}.html'.format(tot_num))
  return print('생성')


# 동별 그래프 출력
def get_graph_dongs(gu, dong_list, start_date, end_date, tot_num) : 
  get_stat_dong_list = []
  for i in range(len(dong_list)) :
    get_stat_dong(gu, dong_list[i], start_date, end_date)
    get_stat_dong_list.append(df_output)

  if len(get_stat_dong_list) > 1 :     
    #############################################
    trace_list = []
    for i in range(len(get_stat_dong_list)) : 
      trace_list.append(go.Scatter(x = get_stat_dong_list[i][get_stat_dong_list[i].columns[0]], y = get_stat_dong_list[i][get_stat_dong_list[i].columns[1]],  name = dong_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))
    fig.update_layout(title = {'text': "<동별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원건수(건/월)")
    fig.write_html('templates/graph_{}.html'.format(tot_num))
    #############################################
    df_merge_pre = get_stat_dongs(gu, dong_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '날짜').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(x ='날짜', figsize = (18,12))
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(연/월)')
    ax.title.set_text('<동별 악취민원발생수>')
    plt.savefig('templates/df{}.jpg'.format(tot_num)) 
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num),sheet_name='Data')
    #############################################
    return print('생성')

  if len(get_stat_dong_list) == 1 :
    #############################################
    trace_list = []
    for i in range(len(get_stat_dong_list)) : 
      trace_list.append(go.Bar(x = get_stat_dong_list[i][get_stat_dong_list[i].columns[0]], y = get_stat_dong_list[i][get_stat_dong_list[i].columns[1]],  name = dong_list[i]))
    fig = go.Figure(data = trace_list)
    fig.update_traces(text = trace_list[0]['y'],  textposition='outside')
    fig.update_traces(marker_color='rgb(0,0,0)')
    fig.update_layout(xaxis_tickfont_size=10, yaxis=dict(tickfont_size=10))    
    fig.update_layout(title = {'text': "<동별 악취민원발생수>", 'x':0.5, 'xanchor': 'center', 'yanchor': 'top'}, xaxis_title="기간(연도/월)", yaxis_title="악취민원건수(건/월)")
    fig.write_html('templates/graph_{}.html'.format(tot_num))
    #############################################
    df_merge_pre = get_stat_dongs(gu, dong_list, start_date, end_date, tot_num)
    df_merge = df_output.sort_values(by = '날짜').fillna(0).reset_index(drop=True)
    ax = df_merge.plot(kind = 'bar', color = 'black', x ='날짜', figsize = (18,12), rot= 45)
    ax.set_xlabel('기간(연도/월)')
    ax.set_ylabel('악취민원 건수(연/월)')
    ax.title.set_text('<동별 악취민원발생수>')
    plt.savefig('templates/df{}.jpg'.format(tot_num)) 
    df_output.to_excel('templates/df{}.xlsx'.format(tot_num), sheet_name='Data')
    #############################################
    return print('생성')


# 동별, 히트맵 출력
def get_map_dong(gu, dong, start_date, end_date, tot_num) : 
  geo_data = "seoul_gu.json"
  with open(geo_data) as f:
      geo_data = json.loads(f.read()) 
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if dong == '전체' :    
    df_fin = df_fin[df_fin['유형_소'] == gu].reset_index()
    latlon = []
    for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
    m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 13, tiles = 'Cartodb Positron')
    HeatMap(latlon, name = 'Heat Map').add_to(m)
    
    # 레이어 추가
    def piste_style_function(feature):
        return {'opacity': 1, 'weight': 1}
    folium.GeoJson(geo_data, style_function = piste_style_function, name='Boundary').add_to(m)
   
    # 버튼 추가
    folium.LayerControl().add_to(m)      
   
    # 범례
    linear = cmp.LinearColormap(['blue', 'green', 'yellow', 'red'], caption='Color Scale for Heat Map')
    linear.add_to(m)
    m.save('templates/heatmap{}.html'.format(tot_num))
    return print("생성")  

  else :
    df_fin = df_fin[(df_fin['유형_소'] == gu)&(df_fin['동_수정'] == dong)].reset_index()
    if len(df_fin[df_fin["동_수정"] == dong]) > 0 :
      latlon = []
      for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
      m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 15, tiles = 'Cartodb Positron')
      HeatMap(latlon, name = 'Heat Map').add_to(m)
      
      # 레이어 추가
      def piste_style_function(feature):
          return {'opacity': 1, 'weight': 1}
      folium.GeoJson(geo_data, style_function = piste_style_function, name='Boundary').add_to(m)
      
      # 버튼 추가
      folium.LayerControl().add_to(m)      
      
      # 범례
      linear = cmp.LinearColormap(['blue', 'green', 'yellow', 'red'], caption='Color Scale for Heat Map')
      linear.add_to(m)
      m.save('templates/heatmap{}.html'.format(tot_num))
      return print("생성")  
    else :
      return print("값없음")  


# 동별, 포인트맵 출력
def get_map_point_dong(gu, dong, start_date, end_date, tot_num) : 
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))]
  if dong == '전체' :    
    df_fin = df_fin[df_fin['유형_소'] == gu].reset_index()
    latlon = []
    for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
    m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 13, tiles = 'Cartodb Positron')
    for i in range(len(latlon)) :
      folium.Circle(location = latlon[i], radius = 50, color = 'Blue', fill_color="#ccc", fill_opacity=1, weight=1, opacity=1, tooltip = df_fin.loc[i,'위반장소']).add_to(m) 
    m.save('templates/mappoint{}.html'.format(tot_num))
    return print("생성")

  else :
    df_fin = df_fin[(df_fin['유형_소'] == gu)&(df_fin['동_수정'] == dong)].reset_index()
    if len(df_fin[df_fin["동_수정"] == dong]) > 0 :
      latlon = []
      for i in range(len(df_fin)) :
        loca_detail = [df_fin['위도'][i] , df_fin['경도'][i]]
        latlon.append(loca_detail)
      m = folium.Map(location = [df_fin['위도'].mean(),  df_fin['경도'].mean()], zoom_start = 15, tiles = 'Cartodb Positron', )
      for i in range(len(latlon)) :
        folium.Circle(location = latlon[i], radius = 50, color = 'Blue', fill_color="#ccc", fill_opacity=1, weight=1, opacity=1, tooltip = df_fin.loc[i,'위반장소']).add_to(m) 
      m.save('templates/mappoint{}.html'.format(tot_num))
      return print("생성")
    else :
      return print("값없음")  


# 전체 동 지도 비교
def get_map_compare_dong(gu, start_date, end_date, tot_num) :  
  df_fin = df[(df['datetime'] >= start_date)&(df['datetime'] <= dt.datetime.strptime(end_date, '%Y-%m') + relativedelta(months=1))][df['유형_소'] == gu]
  df_output = pd.DataFrame(df_fin.groupby(df_fin['동_수정']).count()).reset_index()[['동_수정','번호']]
  df_output.columns = ['동', '민원수(건)']
  
  start_date = start_date.split("-")[0] + '년 ' + start_date.split("-")[1] + '월'
  end_date = end_date.split("-")[0] + '년 ' + end_date.split("-")[1] + '월'

  geo_data = "seoul_dong_v0.2.json"
  with open(geo_data) as f:
      data = json.loads(f.read())
  
  data_list = []
  for i in range(len(data['features'])) : 
    data['features'][i]['properties']['자치구명'] = data['features'][i]['properties'].pop("sggnm")
    data['features'][i]['properties']['동명'] = data['features'][i]['properties'].pop("동")

    for j in range(len(df_output)) : 
      if data['features'][i]['properties']['자치구명'] == gu and data['features'][i]['properties']["동명"] == df_output["동"][j] : 
        data['features'][i]['properties']["민원수(건)"] = int(df_output["민원수(건)"][j])
        del data['features'][i]['properties']['OBJECTID']
        del data['features'][i]['properties']['adm_cd']
        del data['features'][i]['properties']['adm_cd2']
        del data['features'][i]['properties']['adm_nm']
        del data['features'][i]['properties']['sgg']
        # del data['features'][i]['properties']['sggnm']
        del data['features'][i]['properties']['sido']
        del data['features'][i]['properties']['sidonm']
        data['features'][i]['properties']["분석기간"] =  start_date + "~" + end_date
        data['features'][i]['properties']["민원수(건)"] = int(df_output["민원수(건)"][j])  
        data_list.append(data['features'][i])
  
  data['features'] = data_list
  token = 'pk.eyJ1IjoiZHVjazk2NjciLCJhIjoiY2thNzRiMXhxMGQ5NTJ0cXB0dGpmZ3RrOSJ9.f45On_lrv6Nm4iO7oCg7nw'
  res, bins = pd.qcut(df_output['민원수(건)'], 3, retbins=True)
  color_breaks = list(bins) 
  color_stops = create_color_stops(color_breaks, colors='BuPu')

  viz = ChoroplethViz(access_token=token, data=data, color_property='민원수(건)', color_stops=color_stops, center=[data['features'][0]["geometry"]["coordinates"][0][0][0][0], data['features'][0]["geometry"]["coordinates"][0][0][0][1]], zoom=9.5, line_width=2)
  viz.create_html('templates/map_{}.html'.format(tot_num))
  return print('생성')  


#### 통계분석 테이블 생성 함수####
def write_to_html_file(df, tot_num, filename='out.html' ):
       '''
       Write an entire dataframe to an HTML file with nice formatting.
       '''

       result = '''
<html>
<head>
<style>

       h2 {
           text-align: center;
           font-family: Helvetica, Arial, sans-serif;
       }
       table { 
           margin-left: auto;
           margin-right: auto;
       }
       table, th, td {
           border: 1px solid black;
           border-collapse: collapse;
       }
       th, td {
           padding: 5px;
           text-align: center;
           font-family: Helvetica, Arial, sans-serif;
           font-size: 50%;
       }
       thead tr {
        background-color: black;
        color: white;
       }
        tbody tr:nth-child(2n) {
        background-color: #dcdcdc;
       }
        tbody tr:nth-child(2n+1) {
        background-color: #bebebe;
       }
       .wide {
           width: 90%; 
       }

</style>
</head>
<body>
       '''
   
       result += df.to_html(classes='wide', escape=False)
       result += '''
</body>
</html>
'''
       with open(filename, 'w') as f:
           f.write(result)
    
    
# 저장하기            
def get_graph_export(gu_list, start_date, end_date, tot_num, dong_list = None) :   
    if '서울시' in gu_list :

        gu_list = list(df_dong_list.columns[1:])
        get_graph_gus(gu_list, start_date, end_date, tot_num)
        get_map_gu('서울시',start_date, end_date, tot_num )
        get_map_point_gu('서울시',start_date, end_date, tot_num )

    elif dong_list != None:
        if (gu_list[0] != "서울시") and (dong_list[0] == "전체") :
            dong_list = list(df_dong_list[gu_list[0]].dropna())
            gu = gu_list[0]
            get_stat_dongs(gu, dong_list, start_date, end_date, tot_num)
            get_graph_dongs(gu, dong_list, start_date, end_date, tot_num)
            get_map_gu(gu_list[0], start_date, end_date, tot_num)
            get_map_point_gu(gu_list[0],start_date, end_date, tot_num )

        elif (gu_list != ['서울시']) and (dong_list != ['전체']) :
            gu = gu_list[0]
            get_stat_dongs(gu, dong_list, start_date, end_date, tot_num)
            get_graph_dongs(gu, dong_list, start_date, end_date, tot_num)
            get_map_dong(gu_list[0], dong_list[0], start_date, end_date, tot_num)
            get_map_point_dong(gu_list[0], dong_list[0], start_date, end_date, tot_num)

    wb = openpyxl.load_workbook("templates/df{}.xlsx".format(tot_num)) # 엑셀 파일 선택
    sheet_last = wb.create_sheet("Graph") # 제일 마지막에 시트 추가
    ws = wb.worksheets[1] # 시트 지정
    img = Image.open("templates/df{}.jpg".format(tot_num)) # 이미지 지정
    img = openpyxl.drawing.image.Image("templates/df{}.jpg".format(tot_num)) # 이미지 지정
    ws.add_image(img,'A1') # 이미지 삽입
    wb.save("templates/df{}.xlsx".format(tot_num)) # 엑셀 저장  
  
  